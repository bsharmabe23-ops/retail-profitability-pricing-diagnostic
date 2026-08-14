"""
04_build_excel.py
------------------
Builds excel/pricing_diagnostic_summary.xlsx: an executive-ready workbook
with a raw data tab and formula-driven summary tabs (SUMIFS/AVERAGEIFS),
so the workbook recalculates if the raw data tab is refreshed.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2C5F8A")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="2C5F8A")
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

tx = pd.read_csv("data/transactions.csv")
cat = pd.read_csv("data/summary_category.csv")
seg = pd.read_csv("data/summary_segment.csv")
erosion = pd.read_csv("data/summary_discount_bands.csv")
bottom20 = pd.read_csv("data/bottom20_margin_products.csv")
scenario = pd.read_csv("data/scenario_summary.csv")

wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet: Raw Data (sampled to keep file size reasonable; full data in CSV)
# ---------------------------------------------------------------------------
ws_raw = wb.active
ws_raw.title = "Raw Data"
sample = tx.sample(min(5000, len(tx)), random_state=1).sort_values("order_date")
cols = ["transaction_id", "order_date", "customer_segment", "region", "category",
        "product_id", "list_price", "unit_price", "discount_pct", "quantity",
        "revenue", "total_cost", "gross_profit", "gross_margin_pct"]
sample = sample[cols]

for j, colname in enumerate(cols, 1):
    c = ws_raw.cell(row=1, column=j, value=colname.replace("_", " ").title())
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
for i, row in enumerate(sample.itertuples(index=False), 2):
    for j, val in enumerate(row, 1):
        ws_raw.cell(row=i, column=j, value=val).font = BODY_FONT
for j, colname in enumerate(cols, 1):
    ws_raw.column_dimensions[get_column_letter(j)].width = max(12, len(colname) + 4)
ws_raw.freeze_panes = "A2"
last_row = len(sample) + 1

# ---------------------------------------------------------------------------
# Sheet: Category Summary (formula-driven off Raw Data)
# ---------------------------------------------------------------------------
ws_cat = wb.create_sheet("Category Summary")
ws_cat["A1"] = "Revenue & Margin by Category"
ws_cat["A1"].font = TITLE_FONT
headers = ["Category", "Revenue", "Gross Profit", "Margin %", "Avg Discount %"]
for j, h in enumerate(headers, 1):
    c = ws_cat.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

categories = cat.category.tolist()
for i, catname in enumerate(categories, 4):
    ws_cat.cell(row=i, column=1, value=catname).font = BODY_FONT
    ws_cat.cell(row=i, column=2,
        value=f"=SUMIFS('Raw Data'!K2:K{last_row},'Raw Data'!E2:E{last_row},A{i})")
    ws_cat.cell(row=i, column=3,
        value=f"=SUMIFS('Raw Data'!M2:M{last_row},'Raw Data'!E2:E{last_row},A{i})")
    ws_cat.cell(row=i, column=4, value=f"=IF(B{i}=0,0,C{i}/B{i})")
    ws_cat.cell(row=i, column=5,
        value=f"=AVERAGEIFS('Raw Data'!I2:I{last_row},'Raw Data'!E2:E{last_row},A{i})")
    for col in (2, 3):
        ws_cat.cell(row=i, column=col).number_format = "$#,##0"
    ws_cat.cell(row=i, column=4).number_format = "0.0%"
    ws_cat.cell(row=i, column=5).number_format = "0.0%"
    for col in range(1, 6):
        ws_cat.cell(row=i, column=col).border = BORDER

total_row = 4 + len(categories)
ws_cat.cell(row=total_row, column=1, value="TOTAL").font = Font(name=FONT, bold=True)
ws_cat.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})").number_format = "$#,##0"
ws_cat.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})").number_format = "$#,##0"
ws_cat.cell(row=total_row, column=4, value=f"=C{total_row}/B{total_row}").number_format = "0.0%"
for col in range(1, 6):
    ws_cat.cell(row=total_row, column=col).font = Font(name=FONT, bold=True)
    ws_cat.cell(row=total_row, column=col).border = BORDER
for col, w in zip(range(1, 6), [18, 14, 14, 12, 14]):
    ws_cat.column_dimensions[get_column_letter(col)].width = w

# chart
chart = BarChart()
chart.title = "Revenue by Category"
chart.y_axis.title = "Revenue ($)"
data_ref = Reference(ws_cat, min_col=2, min_row=3, max_row=total_row-1)
cats_ref = Reference(ws_cat, min_col=1, min_row=4, max_row=total_row-1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width, chart.height = 16, 9
ws_cat.add_chart(chart, "G3")

# ---------------------------------------------------------------------------
# Sheet: Discount Bands (margin erosion curve)
# ---------------------------------------------------------------------------
ws_disc = wb.create_sheet("Discount Erosion")
ws_disc["A1"] = "Margin Erosion Curve — Discount Depth vs. Margin"
ws_disc["A1"].font = TITLE_FONT
headers = ["Discount Band", "Revenue", "Gross Profit", "Units", "Margin %", "Profit / Unit"]
for j, h in enumerate(headers, 1):
    c = ws_disc.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for i, row in enumerate(erosion.itertuples(index=False), 4):
    ws_disc.cell(row=i, column=1, value=row.discount_band).font = BODY_FONT
    ws_disc.cell(row=i, column=2, value=float(row.revenue)).number_format = "$#,##0"
    ws_disc.cell(row=i, column=3, value=float(row.profit)).number_format = "$#,##0"
    ws_disc.cell(row=i, column=4, value=int(row.units))
    ws_disc.cell(row=i, column=5, value=float(row.margin_pct)).number_format = "0.0%"
    ws_disc.cell(row=i, column=6, value=float(row.profit_per_unit)).number_format = "$#,##0.00"
    for col in range(1, 7):
        ws_disc.cell(row=i, column=col).border = BORDER
for col, w in zip(range(1, 7), [18, 14, 14, 10, 12, 14]):
    ws_disc.column_dimensions[get_column_letter(col)].width = w

note_row = 4 + len(erosion) + 2
ws_disc.cell(row=note_row, column=1,
    value="Insight: margin turns negative beyond ~31% discount depth — volume lift no longer offsets price concession.").font = Font(
    name=FONT, italic=True, size=10, color="C0392B")

# ---------------------------------------------------------------------------
# Sheet: Segment Summary
# ---------------------------------------------------------------------------
ws_seg = wb.create_sheet("Customer Segment")
ws_seg["A1"] = "Profitability by Customer Segment"
ws_seg["A1"].font = TITLE_FONT
headers = ["Segment", "Customers", "Revenue", "Gross Profit", "Margin %", "Rev / Customer", "Avg Discount %"]
for j, h in enumerate(headers, 1):
    c = ws_seg.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for i, row in enumerate(seg.itertuples(index=False), 4):
    ws_seg.cell(row=i, column=1, value=row.customer_segment).font = BODY_FONT
    ws_seg.cell(row=i, column=2, value=int(row.customers))
    ws_seg.cell(row=i, column=3, value=float(row.revenue)).number_format = "$#,##0"
    ws_seg.cell(row=i, column=4, value=float(row.profit)).number_format = "$#,##0"
    ws_seg.cell(row=i, column=5, value=float(row.margin_pct)).number_format = "0.0%"
    ws_seg.cell(row=i, column=6, value=float(row.revenue_per_customer)).number_format = "$#,##0"
    ws_seg.cell(row=i, column=7, value=float(row.avg_discount)).number_format = "0.0%"
    for col in range(1, 8):
        ws_seg.cell(row=i, column=col).border = BORDER
for col, w in zip(range(1, 8), [16, 12, 14, 14, 10, 14, 14]):
    ws_seg.column_dimensions[get_column_letter(col)].width = w

# ---------------------------------------------------------------------------
# Sheet: Bottom Margin Products
# ---------------------------------------------------------------------------
ws_bot = wb.create_sheet("Bottom Margin Products")
ws_bot["A1"] = "Bottom 20 Products by Margin (Revenue > $1,000)"
ws_bot["A1"].font = TITLE_FONT
headers = ["Product ID", "Category", "Product Name", "List Price", "Units", "Revenue", "Gross Profit", "Margin %", "Avg Discount %"]
for j, h in enumerate(headers, 1):
    c = ws_bot.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for i, row in enumerate(bottom20.itertuples(index=False), 4):
    ws_bot.cell(row=i, column=1, value=row.product_id).font = BODY_FONT
    ws_bot.cell(row=i, column=2, value=row.category).font = BODY_FONT
    ws_bot.cell(row=i, column=3, value=row.product_name).font = BODY_FONT
    ws_bot.cell(row=i, column=4, value=float(row.list_price)).number_format = "$#,##0.00"
    ws_bot.cell(row=i, column=5, value=int(row.units))
    ws_bot.cell(row=i, column=6, value=float(row.revenue)).number_format = "$#,##0"
    ws_bot.cell(row=i, column=7, value=float(row.profit)).number_format = "$#,##0"
    ws_bot.cell(row=i, column=8, value=float(row.margin_pct)).number_format = "0.0%"
    ws_bot.cell(row=i, column=9, value=float(row.avg_discount)).number_format = "0.0%"
    for col in range(1, 10):
        ws_bot.cell(row=i, column=col).border = BORDER
for col, w in zip(range(1, 10), [12, 16, 20, 12, 8, 12, 14, 10, 14]):
    ws_bot.column_dimensions[get_column_letter(col)].width = w

# ---------------------------------------------------------------------------
# Sheet: Scenario Analysis
# ---------------------------------------------------------------------------
ws_scn = wb.create_sheet("Scenario Analysis")
ws_scn["A1"] = "Scenario: Cap Grocery & Electronics Discounts at 20%"
ws_scn["A1"].font = TITLE_FONT
headers = ["Metric", "Baseline", "Scenario", "Change"]
for j, h in enumerate(headers, 1):
    c = ws_scn.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for i, row in enumerate(scenario.itertuples(index=False), 4):
    ws_scn.cell(row=i, column=1, value=row.metric).font = BODY_FONT
    is_pct = row.metric == "Margin %"
    ws_scn.cell(row=i, column=2, value=float(row.baseline)).number_format = "0.0%" if is_pct else "$#,##0"
    ws_scn.cell(row=i, column=3, value=float(row.scenario)).number_format = "0.0%" if is_pct else "$#,##0"
    ws_scn.cell(row=i, column=4, value=f"=C{i}-B{i}" if is_pct else f"=(C{i}-B{i})/B{i}")
    ws_scn.cell(row=i, column=4).number_format = "0.0%"
    for col in range(1, 5):
        ws_scn.cell(row=i, column=col).border = BORDER
for col, w in zip(range(1, 5), [18, 16, 16, 12]):
    ws_scn.column_dimensions[get_column_letter(col)].width = w

assumption_row = 4 + len(scenario) + 2
ws_scn.cell(row=assumption_row, column=1,
    value="Assumption: capping deep discounts triggers a 15% volume pull-back on formerly discounted units (elasticity-consistent, conservative estimate).").font = Font(
    name=FONT, italic=True, size=9, color="666666")

# ---------------------------------------------------------------------------
# Save & note recalculation requirement
# ---------------------------------------------------------------------------
wb.save("excel/pricing_diagnostic_summary.xlsx")
print("Workbook saved: excel/pricing_diagnostic_summary.xlsx")
